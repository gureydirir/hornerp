import sys
import os
# Ensure current directory is in python path for Render
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import flet as ft
import shutil
import sqlite3
import datetime
import random # For chart mock data if db is empty
try:
    import receipt_printer
except ImportError:
    receipt_printer = None
    print("Receipt Printer module missing or failed to load.")
try:
    from db_connector import DBHandler, get_pak_time
except ImportError:
    print("\n\nCRITICAL ERROR: 'db_connector.py' is missing!")
    print("You have split the database logic into a new file named 'db_connector.py'.")
    print("Please ensure you UPLOAD 'db_connector.py' to Render alongside 'pos.py'.\n\n")
    raise

# New Import
try:
    from excel_exporter import export_to_excel
except ImportError:
    export_to_excel = None

try:
    from pdf_report import generate_business_report, generate_barcode_sheet
except ImportError:
    print("PDF Report module missing or failed to load.")
    generate_business_report = None
    generate_barcode_sheet = None

# --- THEME COLORS ---
PRIMARY_COLOR = "#1A237E" # Deep Blue
SECONDARY_COLOR = "#00BCD4" # Cyan/Teal
BG_COLOR = "#F5F7FA" 
CARD_BG = "#FFFFFF"
TEXT_COLOR = "#37474F"

# --- TIMEZONE HELPERS ---
def get_pak_time():
    # Pakistan Standard Time is UTC+5
    params = datetime.timezone(datetime.timedelta(hours=5))
    return datetime.datetime.now(params)

class HornERP:
    def __init__(self, page: ft.Page):
        self.page = page
        self.page.title = "Horn ERP - Enterprise Edition"
        self.page.window_icon = "assets/icon.png"
        self.page.theme_mode = ft.ThemeMode.LIGHT
        self.page.bgcolor = BG_COLOR
        self.page.padding = 0
        self.page.window.maximized = True
        
        # Global State
        self.user = {"name": "", "role": "", "id": 0}
        self.customer = {"id": 0, "name": "Walk-in Client", "points": 0}
        self.cart = {}
        self.held_orders = [] # Phase 4: Hold Order
        self.discount = 0 # Phase 4: Discount

        self.init_db()
        self.show_login()

    # --- DATABASE ENGINE ---
    def init_db(self):
        print("🔧 Checking Database...")
        try:
            conn = DBHandler.get_connection()
            cursor = conn.cursor()
            
            auto_id = DBHandler.get_auto_id_sql()
            
            tables = [
                "CREATE TABLE IF NOT EXISTS products (barcode TEXT PRIMARY KEY, name TEXT, price REAL, stock INTEGER DEFAULT 0)",
                f"CREATE TABLE IF NOT EXISTS offline_sales (id {auto_id}, total_amount REAL, cashier_name TEXT, customer_name TEXT, date_created TIMESTAMP DEFAULT CURRENT_TIMESTAMP, is_synced INTEGER DEFAULT 0)",
                f"CREATE TABLE IF NOT EXISTS offline_sale_items (id {auto_id}, sale_id INTEGER, product_name TEXT, price REAL, quantity INTEGER, status TEXT DEFAULT 'sold')",
                f"CREATE TABLE IF NOT EXISTS stock_logs (id {auto_id}, date_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP, product_name TEXT, change_amount INTEGER, reason TEXT, user_role TEXT)",
                "CREATE TABLE IF NOT EXISTS shop_settings (key TEXT PRIMARY KEY, value TEXT)",
                f"CREATE TABLE IF NOT EXISTS expenses (id {auto_id}, date_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP, description TEXT, amount REAL, category TEXT)",
                f"CREATE TABLE IF NOT EXISTS users (id {auto_id}, username TEXT, pin TEXT, role TEXT)",
                f"CREATE TABLE IF NOT EXISTS customers (id {auto_id}, name TEXT, phone TEXT UNIQUE, points INTEGER DEFAULT 0)",
                f"CREATE TABLE IF NOT EXISTS suppliers (id {auto_id}, name TEXT, phone TEXT, address TEXT)"
            ]
            for t in tables: cursor.execute(t)
            conn.commit() # Commit schema creation first
            
            # Defaults - Robust Check
            cursor.execute("SELECT count(*) FROM users")
            user_count = cursor.fetchone()[0]
            print(f"DEBUG: Found {user_count} users in DB.")
            
            if user_count == 0:
                print("DEBUG: Seeding Default Users...")
                cursor.execute("INSERT INTO users (username, pin, role) VALUES ('Admin', '9999', 'Manager')")
                cursor.execute("INSERT INTO users (username, pin, role) VALUES ('Cashier 1', '1234', 'Cashier')")
                conn.commit() # Commit seeds immediately
            
            cursor.execute("SELECT count(*) FROM shop_settings")
            if cursor.fetchone()[0] == 0:
                defaults = [('shop_name', 'HORN ERP'), ('address', 'Mogadishu'), ('phone', '615000000'), ('currency', '$')]
                for k, v in defaults: cursor.execute("INSERT INTO shop_settings VALUES (?, ?)", (k, v))
                conn.commit()

            cursor.execute("SELECT count(*) FROM products")
            if cursor.fetchone()[0] == 0:
                cursor.execute("INSERT INTO products VALUES ('111', 'Test Rice', 1.50, 100)")
                cursor.execute("INSERT INTO products VALUES ('222', 'Test Sugar', 0.90, 50)")
                conn.commit() # Commit seeds

            conn.close()
        except Exception as e:
            print(f"DB Error: {e}")

    def get_setting(self, key):
        try:
            conn = DBHandler.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM shop_settings WHERE key=?", (key,))
            res = cursor.fetchone()
            conn.close()
            return res[0] if res else ""
        except: return ""
        
    def set_setting(self, key, value):
        try:
            conn = DBHandler.get_connection()
            cursor = conn.cursor()
            cursor.execute("INSERT OR REPLACE INTO shop_settings (key, value) VALUES (?, ?)", (key, value))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Setting Error: {e}")
            return False

    def log_stock_change(self, p_name, amount, reason, cursor=None):
        try:
            if cursor:
                user_name = self.user["name"] if self.user["name"] else "Unknown"
                now = get_pak_time().strftime("%Y-%m-%d %H:%M:%S")
                cursor.execute("INSERT INTO stock_logs (date_time, product_name, change_amount, reason, user_role) VALUES (?, ?, ?, ?, ?)", (now, p_name, amount, reason, user_name))
            else:
                conn = DBHandler.get_connection()
                cur = conn.cursor()
                user_name = self.user["name"] if self.user["name"] else "Unknown"
                now = get_pak_time().strftime("%Y-%m-%d %H:%M:%S")
                cur.execute("INSERT INTO stock_logs (date_time, product_name, change_amount, reason, user_role) VALUES (?, ?, ?, ?, ?)", (now, p_name, amount, reason, user_name))
                conn.commit()
                conn.close()
        except Exception as e: print(f"Log Error: {e}")

    def show_snack(self, message, color="green"):
        self.page.snack_bar = ft.SnackBar(ft.Text(message), bgcolor=color)
        self.page.snack_bar.open = True
        self.page.update()

    # --- UI COMPONENTS ---
    def set_app_bar(self, title, show_back_button=False):
        actions = [
            ft.Container(
                content=ft.Row([
                    ft.Icon("person", color="white"),
                    ft.Text(f"{self.user['name']} ({self.user['role']})", color="white", weight=ft.FontWeight.BOLD)
                ]),
                padding=ft.padding.only(right=20)
            )
        ]
        
        if show_back_button:
            back_func = lambda e: self.show_dashboard() if self.user['role'] == 'Manager' else self.show_pos()
            leading = ft.IconButton(icon="arrow_back", icon_color="white", on_click=back_func)
        else:
            leading = ft.Icon(name="store", color="white")

        self.page.appbar = ft.AppBar(
            leading=leading,
            leading_width=60,
            title=ft.Text(title, weight="bold", color="white"),
            center_title=True,
            bgcolor=PRIMARY_COLOR,
            actions=actions,
            elevation=4
        )
        self.page.update()

    def build_kpi_card(self, title, value, icon, gradient_colors):
        return ft.Container(
            content=ft.Row([
                ft.Container(
                    content=ft.Icon(icon, size=40, color="white"),
                    padding=15,
                    bgcolor="#33FFFFFF", # Transparent White
                    border_radius=50,
                ),
                ft.Column([
                    ft.Text(title, color="white", size=16, weight=ft.FontWeight.W_500),
                    ft.Text(value, color="white", size=28, weight=ft.FontWeight.BOLD)
                ], spacing=5)
            ], alignment=ft.MainAxisAlignment.START, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            gradient=ft.LinearGradient(begin=ft.alignment.top_left, end=ft.alignment.bottom_right, colors=gradient_colors),
            border_radius=15,
            padding=20,
            shadow=ft.BoxShadow(blur_radius=15, color="#66000000"),
            expand=True,
            height=130
        )

    def build_section_card(self, title, content_control, icon=None):
        header = ft.Row([
            ft.Icon(name=icon, color=PRIMARY_COLOR) if icon else ft.Container(),
            ft.Text(title, size=20, weight=ft.FontWeight.BOLD, color=PRIMARY_COLOR)
        ], spacing=10)
        
        return ft.Container(
            content=ft.Column([
                header,
                ft.Divider(color="grey300", thickness=1),
                content_control
            ], spacing=15),
            bgcolor=CARD_BG,
            padding=25,
            border_radius=15,
            shadow=ft.BoxShadow(blur_radius=10, color="#1A000000"),
            expand=True
        )

    def build_action_button(self, text, icon, on_click, color=PRIMARY_COLOR, bgcolor="white"):
        return ft.ElevatedButton(
            content=ft.Row([
                ft.Icon(name=icon, color=color),
                ft.Text(text, weight=ft.FontWeight.W_600, color=color)
            ], spacing=10, alignment=ft.MainAxisAlignment.CENTER),
            style=ft.ButtonStyle(
                bgcolor=bgcolor,
                shape=ft.RoundedRectangleBorder(radius=10),
                elevation=2,
                padding=15
            ),
            on_click=on_click,
            width=300
        )

    # --- SCREENS ---
    
    # 1. LOGIN
    def show_login(self):
        self.page.clean()
        self.page.appbar = None
        
        # Ensure attendance table
        try:
            c = DBHandler.get_connection()
            auto_id = DBHandler.get_auto_id_sql()
            c.execute(f"CREATE TABLE IF NOT EXISTS attendance (id {auto_id}, user_id INTEGER, user_name TEXT, action TEXT, timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
            c.close()
        except: pass

        def login_action(e):
            conn = DBHandler.get_connection()
            if not conn:
                error_txt.value = "❌ DB Connection Failed!"
                print("Error: Could not connect to database in login_action")
                self.page.update()
                return

            try:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM users WHERE pin = ?", (pin.value,))
                user_rec = cursor.fetchone()
                conn.close()
            except Exception as e:
                print(f"Login Query Error: {e}")
                error_txt.value = "❌ Database Query Error"
                self.page.update()
                return
            
            if user_rec:
                self.user["id"] = user_rec[0]
                self.user["name"] = user_rec[1]
                self.user["role"] = user_rec[3]
                self.cart = {}
                self.customer = {"id": 0, "name": "Walk-in Client", "points": 0}
                
                if user_rec[3] == "Manager":
                    self.show_dashboard()
                else:
                    self.show_pos()
            else:
                # --- FAILSAFE FOR RENDER/CLOUD ---
                # If database failed to seed, but user enters master pin '9999',
                # auto-create Admin and log in.
                if pin.value == "9999":
                    print("⚠️ Failsafe: Creating Admin user on fly.")
                    try:
                        conn = DBHandler.get_connection()
                        cursor = conn.cursor()
                        # Double check to avoid duplicates if query just failed mysteriously
                        cursor.execute("SELECT * FROM users WHERE pin='9999'")
                        if not cursor.fetchone():
                            cursor.execute("INSERT INTO users (username, pin, role) VALUES ('Admin', '9999', 'Manager')")
                            conn.commit()
                        conn.close()
                        
                        # Set session manually
                        self.user["name"] = "Admin"
                        self.user["role"] = "Manager"
                        self.show_dashboard()
                        self.show_snack("✅ Recovery Mode: Admin Created & Logged In")
                        return
                    except Exception as rec_ex:
                        print(f"Recovery Failed: {rec_ex}")

                error_txt.value = "❌ Invalid PIN"
                self.page.update()

        def clock_action(e):
            conn = DBHandler.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE pin = ?", (pin.value,))
            user_rec = cursor.fetchone()
            
            if not user_rec:
                error_txt.value = "❌ Invalid PIN for Attendance"
                self.page.update()
                conn.close()
                return
            
            # Check last status
            uid, uname = user_rec[0], user_rec[1]
            cursor.execute("SELECT action FROM attendance WHERE user_id=? ORDER BY id DESC LIMIT 1", (uid,))
            last = cursor.fetchone()
            
            new_action = "IN"
            if last and last[0] == "IN": new_action = "OUT"
            
            now_str = datetime.datetime.now().strftime("%Y-mm-%d %H:%M:%S")
            cursor.execute("INSERT INTO attendance (user_id, user_name, action) VALUES (?, ?, ?)", (uid, uname, new_action))
            conn.commit()
            conn.close()
            
            msg = f"✅ {uname}: Clocked {new_action}"
            color = "green" if new_action == "IN" else "orange"
            self.show_snack(msg, color)
            pin.value = ""
            error_txt.value = ""
            self.page.update()
        
        pin = ft.TextField(label="Enter PIN", password=True, width=200, text_align="center", on_submit=login_action, text_size=18, border_color=PRIMARY_COLOR)
        error_txt = ft.Text("", color="red", size=14)
        
        login_card = ft.Container(
            content=ft.Column([
                ft.Icon(name="lock_outline", size=60, color=PRIMARY_COLOR),
                ft.Text("HORN ERP", size=34, weight=ft.FontWeight.BOLD, color=PRIMARY_COLOR),
                ft.Text("Enterprise Login", size=16, color="grey"),
                ft.Divider(height=30, color="transparent"),
                pin,
                ft.Row([
                    ft.ElevatedButton("LOGIN", on_click=login_action, height=50, expand=True, bgcolor=PRIMARY_COLOR, color="white", style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10))),
                    ft.FilledTonalButton("CLOCK IN/OUT", on_click=clock_action, height=50, expand=True, style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10)))
                ], spacing=10, width=300),
                error_txt
            ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=15),
            bgcolor=CARD_BG, padding=40, border_radius=20, shadow=ft.BoxShadow(blur_radius=20, color="#331A237E")
        )
        self.page.add(ft.Container(content=login_card, alignment=ft.alignment.center, expand=True))

    # 2. DASHBOARD
    def show_dashboard(self, period="Daily"):
        try:
            self.page.clean()
            # Disable default App Bar for custom Sidebar layout
            self.page.appbar = None 
            
            shop_name = self.get_setting('shop_name')
            currency = self.get_setting('currency')
            
            conn = DBHandler.get_connection()
            cursor = conn.cursor()
            
            # --- DATE LOGIC ---
            # Python-side time (Pakistan Time)
            now = get_pak_time()
            today_str = now.strftime("%Y-%m-%d")
            month_str = now.strftime("%Y-%m")
            
            # Filter Clauses
            if period == "Daily":
                # CAST to TEXT for Postgres compatibility with LIKE
                where_clause = f"CAST(date_created AS TEXT) LIKE '{today_str}%'"
                chart_limit = 7 # Just show last 7 days context anyway
            elif period == "Weekly":
                # Last 7 Days
                start_date = (now - datetime.timedelta(days=7)).strftime("%Y-%m-%d")
                where_clause = f"date(date_created) >= '{start_date}'"
            elif period == "Monthly":
                where_clause = f"CAST(date_created AS TEXT) LIKE '{month_str}%'"
            
            # --- DATA FETCHING ---
            
            # 1. Total Revenue (Filtered by Period)
            cursor.execute(f"SELECT SUM(total_amount) FROM offline_sales WHERE {where_clause}")
            res = cursor.fetchone()
            revenue_amt = res[0] if res and res[0] else 0.0
            
            # 2. Sales Count (Filtered by Period)
            cursor.execute(f"SELECT COUNT(*) FROM offline_sales WHERE {where_clause}")
            sales_count = cursor.fetchone()[0] or 0
            
            # 3. Total Customers (Global)
            cursor.execute("SELECT COUNT(*) FROM customers")
            total_customers = cursor.fetchone()[0] or 0
            
            # 4. Top Sellers (Filtered by Period)
            # Need join or subquery if we want period-specific top items.
            # For simplicity, we filter the items that belong to sales in that period.
            cursor.execute(f"""
                SELECT i.product_name, SUM(i.quantity) as qty 
                FROM offline_sale_items i
                JOIN offline_sales s ON i.sale_id = s.id
                WHERE {where_clause.replace('date_created', 's.date_created')}
                GROUP BY i.product_name 
                ORDER BY qty DESC LIMIT 5
            """)
            top_sellers = cursor.fetchall()
            
            cursor.execute("SELECT date_time, product_name, change_amount, user_role, reason FROM stock_logs ORDER BY id DESC LIMIT 6")
            recent_logs = cursor.fetchall()
            
            cursor.execute("SELECT name, stock FROM products WHERE stock < 10 ORDER BY stock ASC LIMIT 5")
            low_stock_items = cursor.fetchall()

            # --- ANALYTICS QUERIES (Filtered) ---
            # 1. Employee Performance
            cursor.execute(f"SELECT s.cashier_name, SUM(s.total_amount) FROM offline_sales s WHERE {where_clause} GROUP BY s.cashier_name")
            employee_stats = cursor.fetchall()

            # 2. Customer Insights
            cursor.execute(f"SELECT s.customer_name, SUM(s.total_amount) FROM offline_sales s WHERE {where_clause} GROUP BY s.customer_name ORDER BY SUM(s.total_amount) DESC LIMIT 5")
            customer_stats = cursor.fetchall()

            # 3. Product Revenue
            cursor.execute(f"""
                SELECT i.product_name, SUM(i.price * i.quantity) as revenue 
                FROM offline_sale_items i
                JOIN offline_sales s ON i.sale_id = s.id
                WHERE {where_clause.replace('date_created', 's.date_created')}
                GROUP BY i.product_name 
                ORDER BY revenue DESC LIMIT 5
            """)
            product_revenue_stats = cursor.fetchall()
            
            # 4. Expiry Guard (7 Days)
            try:
                cursor.execute("SELECT name, expiry_date FROM products WHERE expiry_date IS NOT NULL AND expiry_date != '' AND date(expiry_date) BETWEEN date('now') AND date('now', '+7 days')")
                expiring_items = cursor.fetchall()
            except: expiring_items = [] 
            
            conn.close()

            # --- HANDLERS ---
            def run_report_wrapper(e):
                if not generate_business_report:
                    self.show_snack("❌ PDF Module not available", "red")
                    return
                try:
                     # Pass current period to PDF generator
                     filename = generate_business_report(shop_name, period) 
                     if filename.startswith("Error"): 
                         self.show_snack(f"❌ {filename}", color="red")
                     else: 
                         # Launch relative URL
                         self.page.launch_url(f"/reports/{filename}")
                         self.show_snack(f"✅ Report Generated: {filename}")
                except Exception as ex: self.show_snack(f"❌ Error: {ex}", color="red")

            def run_excel_wrapper(e):
                if not export_to_excel: self.show_snack(f"❌ Excel Error: module missing", color="red"); return
                try:
                    filename = export_to_excel(period)
                    if not filename: self.show_snack(f"❌ Error: No file", "red")
                    else: 
                        self.page.launch_url(f"/reports/{filename}")
                        self.show_snack(f"✅ Excel Exported")
                except Exception as ex: self.show_snack(f"❌ Error: {ex}", color="red")
            
            def backup_db(e):
                # ... backup logic ...
                # Backup logic needs similar treatment for Download vs Local Copy
                # But for now user asked about Reports/Excel/Receipts
                try:
                    # Creating a cloud backup mechanism is complex (need to zip logic)
                    # For now just warn
                     self.show_snack("⚠️ Cloud Backup requires Database Dump. Contact Admin.", "orange")
                except Exception as ex: pass


            # --- SIDEBAR COMPONENT ---
            def sidebar_btn(text, icon, func, color="white"):
                return ft.Container(
                    content=ft.Row([
                        ft.Icon(icon, color=color, size=20),
                        ft.Text(text, color=color, size=14, weight=ft.FontWeight.W_500)
                    ], spacing=15),
                    padding=ft.padding.symmetric(vertical=15, horizontal=20),
                    ink=True,
                    on_click=func,
                    border_radius=8
                )

            sidebar = ft.Container(
                content=ft.Column([
                    ft.Container(
                        content=ft.Column([
                            ft.Icon("store_mall_directory", size=40, color="white"),
                            ft.Text("HORN ERP", size=20, weight=ft.FontWeight.BOLD, color="white"),
                            ft.Text(f"{self.user['name']}", size=12, color="white70")
                        ], spacing=5),
                        padding=ft.padding.only(bottom=20)
                    ),
                    ft.Divider(color="white24"),
                    sidebar_btn("Inventory", "inventory_2", lambda e: self.show_inventory()),
                    sidebar_btn("POS Terminal", "point_of_sale", lambda e: self.show_pos()),
                    sidebar_btn("CRM & Clients", "perm_contact_calendar", lambda e: self.show_crm()),
                    sidebar_btn("Staff Manager", "people", lambda e: self.show_staff_screen()),
                    sidebar_btn("Settings", "settings", lambda e: self.show_settings_screen()),
                    ft.Divider(color="white24"),
                    sidebar_btn("Label Maker", "print", lambda e: self.show_label_maker(), color="#AB47BC"),
                    sidebar_btn("PDF Report", "picture_as_pdf", run_report_wrapper, color=SECONDARY_COLOR),
                    sidebar_btn("Excel Export", "table_view", run_excel_wrapper, color="#69F0AE"),
                    sidebar_btn("Backup Data", "backup", backup_db, color="#FFD740"), # Amber for Backup
                    ft.Container(expand=True), # Spacer
                    sidebar_btn("Logout", "logout", lambda e: self.show_login(), color="red100")
                ], spacing=5, scroll="auto", expand=True),
                bgcolor=PRIMARY_COLOR,
                width=250,
                padding=20,
                alignment=ft.alignment.top_left
            )

            # --- MAIN CONTENT COMPONENTS ---
            
            # Period Filter
            def on_period_change(e):
                self.show_dashboard(period=period_dd.value)

            period_dd = ft.Dropdown(
                width=120, text_size=12,
                options=[ft.dropdown.Option("Daily"), ft.dropdown.Option("Weekly"), ft.dropdown.Option("Monthly")],
                value=period, content_padding=5,
                border_color="transparent", bgcolor="white",
                on_change=on_period_change
            )

            # Helper for Charts
            def create_simple_bar_chart(data, color):
                # data is list of (label, value)
                if not data: return ft.Text("No data available", color="grey", size=12)
                max_val = max([x[1] for x in data]) if data else 1
                bars = []
                for label, value in data:
                    pct = (value / max_val) if max_val > 0 else 0
                    bars.append(ft.Column([
                        ft.Row([ft.Text(label, size=12, weight="bold", expand=True), ft.Text(f"{value:,.0f}", size=12)], alignment="spaceBetween"),
                        ft.ProgressBar(value=pct, color=color, bgcolor="grey100", height=6, border_radius=3)
                    ], spacing=2))
                return ft.Column(bars, spacing=8)

            # 1. Top Sellers (Qty)
            top_sellers_ui = create_simple_bar_chart(top_sellers, SECONDARY_COLOR)

            # 2. Employee Performance (Revenue)
            emp_perf_ui = create_simple_bar_chart(employee_stats, "#7B1FA2") # Purple

            # 3. Customer Insights (Revenue)
            cust_perf_ui = create_simple_bar_chart(customer_stats, "#E91E63") # Pink

             # 4. Product Revenue (Revenue)
            prod_rev_ui = create_simple_bar_chart([(x[0], x[1]) for x in product_revenue_stats], "#4CAF50") # Green

            # 5. Activity Feed
            activity_ui = ft.Column(spacing=10)
            if not recent_logs: activity_ui.controls.append(ft.Text("No activity.", size=12, color="grey"))
            else:
                for log in recent_logs:
                    icon = "add" if log[2] > 0 else "remove"
                    color = "green" if log[2] > 0 else "red"
                    activity_ui.controls.append(ft.Row([
                        ft.Icon(icon, size=14, color=color),
                        ft.Text(f"{log[3]}: {log[2]:+}", size=12, weight="bold"),
                        ft.Text(f"{log[1]}", size=12, expand=True, overflow="ellipsis", color="grey")
                    ], spacing=5))

            # 6. Stock Alerts
            stock_alerts_ui = ft.Column(spacing=5)
            if not low_stock_items: stock_alerts_ui.controls.append(ft.Text("Stock healthy.", color="green", size=12))
            else:
                for item in low_stock_items:
                    stock_alerts_ui.controls.append(ft.Row([
                        ft.Icon("warning", color="red", size=14),
                        ft.Text(f"{item[0]} ({item[1]})", color="red", size=12)
                    ]))

            # 7. Expiry Alerts
            expiry_ui = ft.Column(spacing=5)
            if not expiring_items: expiry_ui.controls.append(ft.Text("No expiring items.", color="green", size=12))
            else:
                for item in expiring_items:
                    expiry_ui.controls.append(ft.Row([
                        ft.Icon("dangerous", color="red", size=14),
                        ft.Text(f"{item[0]} ({item[1]})", color="red", size=12, weight="bold")
                    ]))

            # 8. Chart Logic (Simplified for now - Just shows trend)
            # Typically this chart would update to show Weekly trend vs Monthly trend
            # For now we keep the weekly trend visual but metrics are filtered
            revenue_data = [
                ft.LineChartData(
                    data_points=[
                        ft.LineChartDataPoint(0, 100, tooltip="Day 1", point=True),
                        ft.LineChartDataPoint(1, 150, tooltip="Day 2", point=True),
                        ft.LineChartDataPoint(2, 80, tooltip="Day 3", point=True),
                        ft.LineChartDataPoint(3, 220, tooltip="Day 4", point=True),
                        ft.LineChartDataPoint(4, 180, tooltip="Day 5", point=True),
                        ft.LineChartDataPoint(5, 300, tooltip="Day 6", point=True),
                        ft.LineChartDataPoint(6, revenue_amt, tooltip="Current", point=True)
                    ],
                    stroke_width=4,
                    color=SECONDARY_COLOR,
                    stroke_cap_round=True,
                    below_line_bgcolor="#3300BCD4"
                )
            ]
            revenue_chart = ft.LineChart(
                data_series=revenue_data,
                min_y=0, max_y=revenue_amt + 200, height=200, expand=True,
                border=ft.border.all(1, "transparent"),
                left_axis=ft.ChartAxis(show_labels=False),
                bottom_axis=ft.ChartAxis(show_labels=False)
            )

            # Assemble Main Area
            main_content = ft.Container(
                content=ft.Column([
                    # Header
                    ft.Row([
                        ft.Text("Dashboard Overview", size=24, weight="bold", color=TEXT_COLOR),
                        ft.Container(content=ft.Row([ft.Icon("calendar_today", size=16), period_dd], spacing=5), 
                                     bgcolor="white", padding=5, border_radius=8)
                    ], alignment="spaceBetween"),
                    
                    ft.Container(height=10),

                    # KPI Cards
                    ft.Row([
                        self.build_kpi_card("Revenue", f"{currency}{revenue_amt:,.2f}", "attach_money", ["#304FFE", "#00BCD4"]),
                        self.build_kpi_card("Sales", str(sales_count), "receipt_long", ["#7B1FA2", "#E91E63"]),
                        self.build_kpi_card("Customers", str(total_customers), "groups", ["#00C853", "#69F0AE"]),
                    ], spacing=20),

                    ft.Container(height=20),

                    # Chart Section
                    self.build_section_card(f"{period} Performance", revenue_chart, "show_chart"),

                    ft.Container(height=20),
                    
                    ft.Text("Analytics & Insights", size=20, weight="bold", color=PRIMARY_COLOR),
                    
                    # Analytics Grid (2x2)
                    ft.Row([
                        self.build_section_card("Employee Performance", emp_perf_ui, "badge"),
                        self.build_section_card("Customer Insights", cust_perf_ui, "star"),
                    ], spacing=20),
                    
                    ft.Container(height=10),
                    
                    ft.Row([
                         self.build_section_card("Product Revenue", prod_rev_ui, "monetization_on"),
                         self.build_section_card("Top Sellers (Qty)", top_sellers_ui, "leaderboard"),
                    ], spacing=20),

                    ft.Container(height=20),
                    
                    ft.Text("Operational Status", size=20, weight="bold", color=PRIMARY_COLOR),

                    # Operations Grid
                    ft.Row([
                        self.build_section_card("Recent Activity", activity_ui, "history"),
                        self.build_section_card("Low Stock Alerts", stock_alerts_ui, "notification_important"),
                        self.build_section_card("Expiry Guard (7 Days)", expiry_ui, "health_and_safety")
                    ], spacing=20, alignment="start", vertical_alignment="start")

                ], scroll=ft.ScrollMode.AUTO),
                expand=True,
                padding=30
            )

            # Final App Layout
            self.page.add(ft.Row([sidebar, main_content], spacing=0, expand=True))

        except Exception as e:
            import traceback
            traceback.print_exc()
            self.page.clean()
            self.page.add(ft.Text(f"Dashboard Error: {e}", color="red"))
            self.page.update()

    # 3. POS TERMINAL
    def show_pos(self):
        try:
            self.page.clean()
            currency = self.get_setting('currency')
            self.discount = 0 # Reset discount on new session
            
            role_title = f"CASHIER: {self.user['name']}" if self.user['role'] != 'Manager' else "POS Terminal"
            
            # Action Bar
            actions = []
            if self.held_orders:
                actions.append(ft.Container(
                    content=ft.Row([ft.Icon("pause_circle", color="orange"), ft.Text(f"{len(self.held_orders)} Held", color="orange", weight="bold")], spacing=5),
                    padding=10, ink=True, on_click=lambda e: show_held_orders_dialog()
                ))
            actions.append(ft.IconButton(icon="logout", icon_color="white", on_click=lambda e: self.show_login()))

            if self.user["role"] == "Manager": 
                self.set_app_bar(role_title, show_back_button=True)
            else:
                self.page.appbar = ft.AppBar(
                    title=ft.Text(role_title, weight="bold", color="white"), 
                    bgcolor=PRIMARY_COLOR, 
                    actions=actions
                )
                self.page.update()
            
            # Responsive Grid
            def get_grid_runs(width):
                if width < 500: return 2
                if width < 900: return 3
                if width < 1200: return 4
                return 5

            products_grid = ft.GridView(
                expand=True, 
                runs_count=get_grid_runs(self.page.width), 
                child_aspect_ratio=0.85, 
                spacing=10, 
                run_spacing=10
            )
            
            def on_resize(e):
                products_grid.runs_count = get_grid_runs(self.page.width)
                products_grid.update()
                
            self.page.on_resize = on_resize

            cart_list = ft.Column(scroll="auto", expand=True, spacing=5)
            
            # Totals
            subtotal_text = ft.Text(f"{currency}0.00", size=16, color="grey")
            discount_text = ft.Text(f"-{currency}0.00", size=16, color="red")
            total_text = ft.Text(f"{currency}0.00", size=34, weight=ft.FontWeight.BOLD, color=PRIMARY_COLOR)
            
            phone_input = ft.TextField(label="Client Phone", hint_text="Enter number...", text_size=16, height=50, prefix_icon="phone", border_color=PRIMARY_COLOR, expand=True)
            client_name_label = ft.Text("Walk-in Client", weight="bold", color="grey")

            # --- LOGIC ---
            def check_client(e):
                phone = phone_input.value.strip()
                if not phone: return
                conn = DBHandler.get_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM customers WHERE phone=?", (phone,))
                row = cursor.fetchone()
                conn.close()
                
                if row:
                    client_name_label.value = f"Client: {row[1]} (⭐ {row[3]})"
                    client_name_label.color = "blue"
                    self.customer = {"id": row[0], "name": row[1], "points": row[3]}
                    self.show_snack(f"✅ Found Client: {row[1]}")
                else:
                    client_name_label.value = "New Client (Will be saved)"
                    client_name_label.color = "orange"
                    self.customer = {"id": 0, "name": "New Client", "points": 0}
                self.page.update()
            
            # DEFAULT TO WALK-IN (Privacy/Speed)
            self.customer = {"id": 0, "name": "Walk-in", "points": 0} 
            
            # Helper to toggle client input
            def toggle_client_input(e):
                client_input_row.visible = not client_input_row.visible
                self.page.update() 

            def add_to_cart(item):
                barcode = item[0]
                if barcode in self.cart: 
                    self.cart[barcode]['qty'] += 1
                else: 
                    self.cart[barcode] = {'name': item[1], 'price': item[2], 'qty': 1, 'barcode': barcode}
                self.show_snack(f"Added: {item[1]}")
                update_cart()
                
            def filter_products(e):
                render_products(search_term=pos_search.value)

            def update_cart():
                cart_list.controls.clear()
                subtotal = 0
                for barcode, item in self.cart.items():
                    line_total = item['price'] * item['qty']
                    subtotal += line_total
                    cart_list.controls.append(ft.Container(
                        content=ft.Row([
                            ft.Text(f"{item['qty']}x", weight=ft.FontWeight.BOLD, color=PRIMARY_COLOR, width=30),
                            ft.Text(item['name'], expand=True, color=TEXT_COLOR, size=15, weight=ft.FontWeight.W_500),
                            ft.Text(f"{currency}{line_total:.2f}", weight=ft.FontWeight.BOLD, color="black"),
                            ft.IconButton(icon="delete_outline", icon_color="red400", on_click=lambda e, b=barcode: remove_item(b))
                        ]), padding=10, bgcolor="white", border_radius=8, shadow=ft.BoxShadow(blur_radius=2, color="grey200")
                    ))
                
                final_total = max(0, subtotal - self.discount)
                subtotal_text.value = f"Subtotal: {currency}{subtotal:.2f}"
                discount_text.value = f"Discount: -{currency}{self.discount:.2f}"
                total_text.value = f"{currency}{final_total:.2f}"
                self.page.update()

            def remove_item(barcode):
                if barcode in self.cart:
                    del self.cart[barcode]
                    update_cart()

            # --- FEATURES: HOLD / RECALL / DISCOUNT ---
            def hold_order(e):
                if not self.cart: return
                order_data = {
                    "cart": self.cart.copy(),
                    "customer": self.customer.copy(),
                    "time": datetime.datetime.now().strftime("%H:%M:%S")
                }
                self.held_orders.append(order_data)
                self.cart = {}
                self.customer = {"id": 0, "name": "Walk-in Client", "points": 0}
                self.discount = 0
                phone_input.value = ""
                client_name_label.value = "Walk-in Client"
                update_cart()
                self.show_snack("⏸️ Order Held")
                self.show_pos() # Refresh to update appbar counter

            def show_held_orders_dialog():
                def restore_order(idx):
                    data = self.held_orders.pop(idx)
                    self.cart = data["cart"]
                    self.customer = data["customer"]
                    phone_input.value = self.customer.get("phone", "") # Phone might not be stored perfectly in struct logic above, simplified for now
                    self.show_pos()
                    self.show_snack("▶️ Order Resumed")
                    dlg.open = False

                items = []
                for idx, order in enumerate(self.held_orders):
                    total = sum(i['price'] * i['qty'] for i in order['cart'].values())
                    items.append(ft.ListTile(
                        leading=ft.Icon(ft.icons.RECEIPT_LONG),
                        title=ft.Text(f"Held at {order['time']} - {order['customer']['name']}"),
                        subtitle=ft.Text(f"Total: {currency}{total:.2f}"),
                        trailing=ft.IconButton(ft.icons.PLAY_ARROW, on_click=lambda e, i=idx: restore_order(i))
                    ))

                dlg = ft.AlertDialog(title=ft.Text("Held Orders"), content=ft.Column(items, height=300, scroll="auto"))
                if hasattr(self.page, 'open'): self.page.open(dlg)
                else:
                    self.page.dialog = dlg
                    dlg.open = True
                    self.page.update()

            def show_discount_dialog(e):
                self.show_snack("Opening Discount...", "blue")
                disc_input = ft.TextField(label="Discount Amount", keyboard_type="number", autofocus=False)
                def apply_disc(e):
                    if not disc_input.value: return
                    try:
                        val = float(disc_input.value)
                        if val < 0: raise ValueError("Negative")
                        self.discount = val
                        update_cart()
                        
                        if hasattr(self.page, 'close'): self.page.close(dlg)
                        else: dlg.open = False
                        self.page.update()
                        
                        self.show_snack(f"Discount Applied: {currency}{val}")
                    except ValueError:
                        disc_input.error_text = "Invalid Number"
                        disc_input.update()
                
                dlg = ft.AlertDialog(title=ft.Text("Apply Discount"), content=disc_input, actions=[ft.TextButton("Apply", on_click=apply_disc)])
                
                if hasattr(self.page, 'open'): self.page.open(dlg)
                else:
                    self.page.dialog = dlg
                    dlg.open = True
                    self.page.update()

            def reprint_last(e):
                if hasattr(self, 'last_receipt_data') and self.last_receipt_data:
                    filename = receipt_printer.print_receipt(self.last_receipt_data)
                    if filename: self.page.launch_url(f"/reports/{filename}")
                    self.show_snack("Reprinting last receipt...", "blue")
                else:
                    self.show_snack("No receipt found to reprint", "orange")

            def show_payment_dialog(e):
                self.show_snack("Processing Payment...", "blue")
                
                if not self.cart: 
                    self.show_snack("Cart is empty!", "red")
                    return
                
                # 1. Calculation
                subtotal = sum(item['price'] * item['qty'] for item in self.cart.values())
                total_due = max(0, subtotal - self.discount)
                
                # 2. Controls
                cash_input = ft.TextField(
                    label="Cash Amount", 
                    value=str(total_due), # Default to exact amount
                    text_size=24, 
                    # weight="bold", # REMOVED: Invalid arg
                    suffix_text=currency, 
                    border_color=PRIMARY_COLOR,
                    keyboard_type="number",
                    autofocus=False 
                )
                
                change_label = ft.Text(f"Change: {currency}0.00", size=18, color="green")
                
                pay_btn = ft.ElevatedButton(
                    "CONFIRM PAYMENT", 
                    bgcolor="green", 
                    color="white", 
                    height=50, 
                    width=250
                )

                def on_cash_change(e):
                    try:
                        val = float(cash_input.value) if cash_input.value else 0
                        change = val - total_due
                        if change >= 0:
                            change_label.value = f"Change: {currency}{change:.2f}"
                            change_label.color = "green"
                            pay_btn.disabled = False
                        else:
                            change_label.value = f"Owe: {currency}{abs(change):.2f}"
                            change_label.color = "red"
                            pay_btn.disabled = False 
                        cash_input.update()
                        change_label.update()
                        pay_btn.update()
                    except: pass
                
                cash_input.on_change = on_cash_change

                def do_pay(e):
                    try:
                        val = float(cash_input.value) if cash_input.value else 0
                        if val < total_due and abs(val - total_due) > 0.01:
                            self.show_snack("Insufficient Cash", "red")
                            return
                        
                        # Success
                        self.last_cash_tendered = val
                        # Close dialog safely
                        if hasattr(self.page, 'close'):
                             self.page.close(dlg)
                        else:
                             dlg.open = False
                             self.page.update()
                             
                        process_payment(None)
                    except ValueError:
                        self.show_snack("Invalid Amount", "red")

                pay_btn.on_click = do_pay
                
                # 3. Simple Vertical Layout
                content_col = ft.Column([
                    ft.Text("Payment Details", size=20, weight="bold", color=PRIMARY_COLOR),
                    ft.Divider(),
                    ft.Row([
                        ft.Text("Total To Pay:", size=16),
                        ft.Text(f"{currency}{total_due:.2f}", size=24, weight="bold")
                    ], alignment="spaceBetween"),
                    ft.Container(height=10),
                    cash_input,
                    ft.Container(height=10),
                    change_label,
                    ft.Container(height=20),
                    pay_btn
                ], tight=True, width=350, spacing=10)

                dlg = ft.AlertDialog(
                    content=ft.Container(content=content_col, padding=20),
                    modal=True
                )
                
                # USE MODERN API
                if hasattr(self.page, 'open'):
                    self.page.open(dlg)
                else:
                    # Fallback for old flet if somehow running there, but unlikely
                    self.page.dialog = dlg
                    dlg.open = True
                    self.page.update()

            def process_payment(e):
                if not self.cart: 
                    self.show_snack("Cart is empty!", "red")
                    return
                
                # Ensure customer set (redundant but safe)
                if not self.customer:
                    self.customer = {"id": 0, "name": "Walk-in", "points": 0}

                try:
                    conn = DBHandler.get_connection()
                    cursor = conn.cursor()
                    
                    # 2. Calculate Totals
                    subtotal = sum(item['price'] * item['qty'] for item in self.cart.values())
                    total = max(0, subtotal - self.discount)
                    pkt_now = get_pak_time()
                    date_str = pkt_now.strftime("%Y-%m-%d %H:%M:%S")

                    # 3. Insert Sale
                    if DBHandler.HAS_POSTGRES:
                        # Postgres requires RETURNING id handling which simpleexecute might not expose easily without SafeCursor helper
                        sale_id = DBHandler.execute_and_get_id(cursor, "INSERT INTO offline_sales (total_amount, cashier_name, customer_name, date_created, is_synced) VALUES (?, ?, ?, ?, 0)", (total, self.user['name'], self.customer['name'], date_str))
                    else:
                        cursor.execute("INSERT INTO offline_sales (total_amount, cashier_name, customer_name, date_created, is_synced) VALUES (?, ?, ?, ?, 0)", 
                                    (total, self.user['name'], self.customer['name'], date_str))
                        sale_id = cursor.lastrowid
                    
                    # 4. Insert Items & Update Stock
                    receipt_items = []
                    for barcode, item in self.cart.items():
                        cursor.execute("UPDATE products SET stock = stock - ? WHERE barcode=?", (item['qty'], barcode))
                        cursor.execute("INSERT INTO offline_sale_items (sale_id, product_name, quantity, price) VALUES (?, ?, ?, ?)", 
                                       (sale_id, item['name'], item['qty'], item['price']))
                        
                        # Log Activity
                        self.log_stock_change(item['name'], -item['qty'], f"Sale #{sale_id}", cursor=cursor)
                        
                        receipt_items.append({
                            "name": item['name'],
                            "qty": item['qty'],
                            "price": item['price'],
                            "total": item['qty'] * item['price']
                        })

                    conn.commit()
                    conn.close()
                    
                    # 5. Print Receipt
                    # Use last_cash_tendered if available, else assume exact
                    cash_tendered = getattr(self, 'last_cash_tendered', total) 
                    change_due = cash_tendered - total

                    receipt_data = {
                        "shop_name": self.get_setting('shop_name'),
                        "address": self.get_setting('address'),
                        "phone": self.get_setting('phone'),
                        "currency": currency,
                        "receipt_no": sale_id,
                        "date": date_str,
                        "cashier": self.user['name'],
                        "customer": self.customer['name'],
                        "items": receipt_items,
                        "subtotal": subtotal,
                        "discount": self.discount,
                        "total": total,
                        "cash": cash_tendered,
                        "change": change_due
                    }
                    
                    try:
                        # Modified to handle Web Print (Launch URL)
                        filename = receipt_printer.print_receipt(receipt_data)
                        if filename:
                            self.page.launch_url(f"/reports/{filename}")
                    except Exception as print_err:
                        print(f"Printing Error (Ignored): {print_err}")
                        self.show_snack(f"⚠️ Sale Saved, but Print Failed: {print_err}", "orange")

                    self.last_receipt_data = receipt_data
                    self.last_receipt_data = receipt_data

                    # 6. Reset UI
                    self.cart = {}
                    self.discount = 0
                    self.customer = {"id": 0, "name": "Walk-in", "points": 0} # Reset to default
                    update_cart()
                    client_name_label.value = "Walk-in Client"
                    client_name_label.color = "grey"
                    
                    self.show_pos() # Refresh UI
                    self.show_snack("✅ Sale Completed!", "green")
                    
                except Exception as ex:
                    print(f"Sale Error: {ex}")
                    self.show_snack(f"Error: {ex}", "red")

            # --- CATEGORY SYSTEM & RENDER ---
            
            selected_category = {"name": "All"}

            def render_products(search_term=""):
                conn = DBHandler.get_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM products")
                rows = cursor.fetchall()
                conn.close()
                
                # Fetch categories for tabs
                # Assume category is index 5. If missing, "General"
                # Need to be safe with indices
                
                products_grid.controls.clear()
                
                current_cat = selected_category["name"]

                for row in rows:
                    p_cat = row[5] if len(row) > 5 and row[5] else "General"
                    p_name = row[1]
                    p_code = row[0]
                    
                    # Filters
                    if current_cat != "All" and p_cat != current_cat: continue
                    if search_term and search_term.lower() not in p_name.lower() and search_term not in p_code: continue

                    # Advanced Card Design
                    color_map = {
                        "General": [PRIMARY_COLOR, "#303F9F"],
                        "Pantry": ["#FF9800", "#F57C00"],        # Orange
                        "Dairy": ["#03A9F4", "#0288D1"],         # Light Blue
                        "Beverages": ["#009688", "#00796B"],     # Teal
                        "Household": ["#607D8B", "#455A64"],     # Blue Grey
                        "Personal Care": ["#9C27B0", "#7B1FA2"], # Purple
                        "Snacks": ["#E91E63", "#C2185B"],        # Pink
                        "Fresh Food": ["#4CAF50", "#2E7D32"],    # Green
                        "Electronics": ["#607D8B", "#455A64"]
                    }
                    bg_gradient = color_map.get(p_cat, [PRIMARY_COLOR, "#303F9F"])
                    
                    # If specific well known categories use them, otherwise hash a color or just default
                    
                    card = ft.Container(
                        content=ft.Column([
                            # Header Image / Icon Area
                            ft.Container(
                                content=ft.Column([
                                    ft.Row([
                                         ft.Container(
                                            content=ft.Text(f"{row[3]}", size=10, color="white", weight="bold"),
                                            bgcolor="red" if row[3] < 10 else "rgba(0,0,0,0.5)",
                                            padding=ft.padding.symmetric(horizontal=6, vertical=2),
                                            border_radius=8
                                        )
                                    ], alignment="end"),
                                    ft.Icon("local_mall", size=40, color="white"),
                                ], alignment="spaceBetween"),
                                height=80,
                                padding=10,

                                gradient=ft.LinearGradient(
                                    begin=ft.alignment.top_left,
                                    end=ft.alignment.bottom_right,
                                    colors=bg_gradient
                                ),
                                border_radius=ft.border_radius.only(top_left=15, top_right=15)
                            ),
                            
                            # Content Area
                            ft.Container(
                                content=ft.Column([
                                    ft.Text(row[1], size=13, weight="bold", max_lines=2, overflow="ellipsis", color=TEXT_COLOR),
                                    ft.Text(p_cat, size=10, color="grey"),
                                    ft.Row([
                                        ft.Text(f"{currency}{row[2]:.2f}", weight="bold", size=16, color=PRIMARY_COLOR),
                                        ft.Container(
                                            content=ft.Icon("add", color="white", size=16),
                                            bgcolor=PRIMARY_COLOR,
                                            shape=ft.BoxShape.CIRCLE,
                                            padding=5,
                                        )
                                    ], alignment="spaceBetween")
                                ], spacing=2),
                                padding=10,
                                bgcolor="white",
                                border_radius=ft.border_radius.only(bottom_left=15, bottom_right=15)
                            )
                        ], spacing=0),
                        
                        border_radius=15,
                        ink=True,
                        on_click=lambda e, r=row: add_to_cart(r),
                        shadow=ft.BoxShadow(blur_radius=10, color="grey300", offset=ft.Offset(0, 5)),
                        tooltip=f"{row[1]} - Stock: {row[3]}"
                    )
                    products_grid.controls.append(card)
                self.page.update()

            pos_search = ft.TextField(hint_text="Search items...", prefix_icon="search", on_change=filter_products, border_radius=10, bgcolor="white", height=40, text_size=14, content_padding=10)

            # Build Categories
            # We need to fetch meaningful categories from DB
            def build_categories():
                conn = DBHandler.get_connection()
                cur = conn.cursor()
                try:
                    cur.execute("SELECT DISTINCT category FROM products")
                    cats = [x[0] for x in cur.fetchall() if x[0]]
                except: cats = []
                conn.close()
                if not cats: cats = ["General"]
                
                full_cats = ["All"] + sorted(cats)
                
                cat_controls = []
                for c in full_cats:
                    is_active = (c == selected_category["name"])
                    cat_controls.append(
                        ft.Container(
                            content=ft.Text(c, color="white" if is_active else PRIMARY_COLOR, weight="bold"),
                            bgcolor=PRIMARY_COLOR if is_active else "white",
                            border=ft.border.all(1, PRIMARY_COLOR),
                            border_radius=20,
                            padding=ft.padding.symmetric(horizontal=15, vertical=8),
                            ink=True,
                            on_click=lambda e, cat=c: change_category(cat)
                        )
                    )
                return ft.Row(cat_controls, scroll="auto")

            def change_category(cat_name):
                selected_category["name"] = cat_name
                # Re-render categories to update styling (active state)
                cat_bar.content = build_categories()
                render_products(pos_search.value)

            cat_bar = ft.Container(content=build_categories(), height=50)

            self.page.add(ft.Row([
                ft.Container(
                    content=ft.Column([
                        ft.Row([pos_search, cat_bar], spacing=10),
                        ft.Container(content=products_grid, expand=True) # Grid expands inside
                    ], spacing=10), 
                    expand=7, padding=20
                ),
                ft.Container(content=ft.Column([
                    ft.Text("Current Order", size=24, weight=ft.FontWeight.BOLD, color=PRIMARY_COLOR),
                    ft.Divider(),
                    ft.Row([client_name_label, ft.IconButton(icon="edit", icon_size=16, tooltip="Change Client", on_click=toggle_client_input)], alignment="spaceBetween"),
                    client_input_row := ft.Row([phone_input, ft.IconButton(icon="check_circle", icon_color="green", tooltip="Check Client", on_click=check_client)], visible=False),
                    ft.Divider(),
                    ft.Container(content=cart_list, expand=True, bgcolor="grey50", border_radius=10, padding=10),
                    ft.Divider(),
                    ft.Column([
                        ft.Row([subtotal_text, discount_text], alignment="spaceBetween"),
                        ft.Row([ft.Text("Total:", size=20), total_text], alignment="spaceBetween"),
                    ]),
                    ft.Row([
                        ft.ElevatedButton("Clear", on_click=lambda e: [self.cart.clear(), update_cart()], bgcolor="red", color="white", expand=True),
                        ft.ElevatedButton("Discount", on_click=show_discount_dialog, bgcolor="orange", color="white", expand=True)
                    ]),
                    ft.ElevatedButton("PROCEED TO PAYMENT", on_click=show_payment_dialog, bgcolor=PRIMARY_COLOR, color="white", height=60, style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10))),
                    ft.TextButton("Reprint Last Receipt", on_click=reprint_last, icon="print")
                ], spacing=15), width=400, bgcolor=CARD_BG, padding=25, border_radius=ft.border_radius.only(top_left=20, bottom_left=20), shadow=ft.BoxShadow(blur_radius=20, color="grey300"))
            ], expand=True, spacing=0))
            render_products()

        except Exception as e:
            import traceback
            traceback.print_exc()
            self.page.clean()
            self.page.add(
                ft.Column([
                    ft.Icon("error", color="red", size=60),
                    ft.Text("POS Error", size=30, color="red", weight="bold"),
                    ft.Text(f"Error details: {e}", size=20, color="black"),
                    ft.ElevatedButton("Go Back", on_click=lambda _: self.show_login())
                ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER)
            )
            self.page.update()

    # --- IMPLEMENTED QUICK ACTIONS ---

    # 4. INVENTORY MANAGEMENT (Replaces Supply Chain)
    def show_inventory(self):
        try:
            self.page.clean()
            self.set_app_bar("Inventory Management", show_back_button=True)
            currency = self.get_setting('currency')
            
            # --- INPUTS ---
            barcode_tf = ft.TextField(label="Barcode", width=150, border_color=PRIMARY_COLOR)
            name_tf = ft.TextField(label="Product Name", expand=True, border_color=PRIMARY_COLOR)
            price_tf = ft.TextField(label="Price", width=100, keyboard_type="number", border_color=PRIMARY_COLOR)
            stock_tf = ft.TextField(label="Stock", width=100, keyboard_type="number", border_color=PRIMARY_COLOR)
            expiry_tf = ft.TextField(label="Expiry (YYYY-MM-DD)", width=160, border_color="red", hint_text="Optional")
            category_tf = ft.TextField(label="Category", width=150, border_color=PRIMARY_COLOR, hint_text="e.g. Food")

            search_tf = ft.TextField(label="Search Inventory...", prefix_icon="search", expand=True, on_change=lambda e: load_products(), border_color=PRIMARY_COLOR)
            
            edit_mode = {"active": False}

            # --- TABLE ---
            products_table = ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text("Barcode")),
                    ft.DataColumn(ft.Text("Product")),
                    ft.DataColumn(ft.Text("Category")),
                    ft.DataColumn(ft.Text("Price"), numeric=True),
                    ft.DataColumn(ft.Text("Stock"), numeric=True),
                    ft.DataColumn(ft.Text("Expiry")),
                    ft.DataColumn(ft.Text("Actions")),
                ],
                heading_row_color="grey200",
                border=ft.border.all(1, "grey300"),
                border_radius=10,
                column_spacing=20,
                width=float("inf") # Expand horizontally
            )

            # --- BULK IMPORT LOGIC ---
            def download_template(e):
                try:
                    import openpyxl
                    desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop')
                    fname = "HornERP_Inventory_Template.xlsx"
                    fp = os.path.join(desktop, fname)
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Import Template"
                    
                    # Headers
                    headers = ["Barcode", "Name", "Category", "Price", "Stock", "Expiry Date"]
                    ws.append(headers)
                    
                    # Example
                    example_expiry = (datetime.datetime.now() + datetime.timedelta(days=90)).strftime("%Y-%m-%d")
                    ws.append(["1001", "Example Rice", "Pantry", 1.50, 100, example_expiry])
                    
                    # Styling
                    for cell in ws[1]:
                        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                        cell.fill = openpyxl.styles.PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                        
                    ws.column_dimensions["B"].width = 25
                    ws.column_dimensions["C"].width = 15
                    
                    wb.save(fp)
                    self.show_snack(f"✅ Template Saved: {fname}", "blue")
                except ImportError:
                    self.show_snack("❌ Error: 'openpyxl' module missing. Please install it.", "red")
                except Exception as ex: self.show_snack(f"Error: {ex}", "red")

            def process_import(e: ft.FilePickerResultEvent):
                if not e.files: return
                
                fpath = e.files[0].path
                ext = fpath.split('.')[-1].lower()
                count = 0
                errors = 0
                
                try:
                    conn = DBHandler.get_connection()
                    cursor = conn.cursor()
                    
                    rows_to_process = []
                    
                    if ext in ["xlsx", "xls"]:
                        import openpyxl
                        wb = openpyxl.load_workbook(fpath, data_only=True) # data_only to get values not formulas
                        ws = wb.active
                        
                        # Assuming Row 1 is headers
                        # Map headers to indices
                        headers = {}
                        for cell in ws[1]:
                            if cell.value:
                                headers[str(cell.value).strip().lower()] = cell.column - 1
                                
                        # Required mapping
                        req_cols = ["barcode", "name", "price", "stock"]
                        if not all(k in headers for k in req_cols):
                            self.show_snack("❌ Invalid Excel. Missing headers (Barcode, Name, Price, Stock).", "red")
                            return

                        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                            try:
                                # Safe extraction
                                def get_val(key):
                                    idx = headers.get(key)
                                    val = row[idx] if idx is not None and idx < len(row) else None
                                    return val

                                r_data = {
                                    "barcode": str(get_val("barcode")),
                                    "name": str(get_val("name")),
                                    "category": str(get_val("category")) if get_val("category") else "General",
                                    "price": float(get_val("price")) if get_val("price") is not None else 0.0,
                                    "stock": int(get_val("stock")) if get_val("stock") is not None else 0,
                                    "expiry_date": str(get_val("expiry date")) if get_val("expiry date") else "" 
                                    # Note: expiry header might differ, try fallback
                                }
                                if not r_data["expiry_date"]:
                                     r_data["expiry_date"] = str(get_val("expiry")) if get_val("expiry") else ""

                                rows_to_process.append(r_data)
                            except: errors += 1

                    elif ext == "csv":
                        with open(fpath, "r") as f:
                            reader = csv.DictReader(f)
                            # Normalize headers
                            # ... (CSV logic if needed, but user focused on Excel)
                             # Headers
                            headers_map = {h.strip().lower(): h for h in reader.fieldnames}
                            
                            for row in reader:
                                try:
                                    # Normalize row keys
                                    row_lower = {k.strip().lower(): v for k,v in row.items()}
                                    
                                    r_data = {
                                        "barcode": row_lower.get("barcode"),
                                        "name": row_lower.get("name"),
                                        "category": row_lower.get("category", "General"),
                                        "price": float(row_lower.get("price", 0)),
                                        "stock": int(row_lower.get("stock", 0)),
                                        "expiry_date": row_lower.get("expiry date", row_lower.get("expiry_date", ""))
                                    }
                                    rows_to_process.append(r_data)
                                except: errors += 1

                    # Process Upserts
                    for item in rows_to_process:
                        try:
                            b, n, c, p, s, ex = item["barcode"], item["name"], item["category"], item["price"], item["stock"], item["expiry_date"]
                            if not b or not n: continue
                            
                            cursor.execute("SELECT barcode FROM products WHERE barcode=?", (b,))
                            if cursor.fetchone():
                                cursor.execute("UPDATE products SET name=?, category=?, price=?, stock=?, expiry_date=? WHERE barcode=?", (n, c, p, s, ex, b))
                                self.log_stock_change(n, 0, "Bulk Update", cursor=cursor)
                            else:
                                cursor.execute("INSERT INTO products (barcode, name, category, price, stock, expiry_date) VALUES (?, ?, ?, ?, ?, ?)", (b, n, c, p, s, ex))
                                self.log_stock_change(n, s, "Bulk Import", cursor=cursor)
                            count += 1
                        except: errors += 1
                            
                    conn.commit()
                    conn.close()
                    load_products()
                    self.show_snack(f"✅ Imported {count} items. {errors} failed/skipped.", "green")
                    
                except Exception as ex:
                    self.show_snack(f"Import Error: {ex}", "red")

            file_picker = ft.FilePicker(on_result=process_import)
            self.page.overlay.append(file_picker)

            # Ensure schema is ready (Migration)
            def ensure_schema():
                try:
                    conn = DBHandler.get_connection()
                    cursor = conn.cursor()
                    # Only try ALTER if it's sqlite, or handle Postgres carefully.
                    # For simplicity, we just try/except.
                    try: cursor.execute("ALTER TABLE products ADD COLUMN expiry_date TEXT")
                    except: pass
                    try: cursor.execute("ALTER TABLE products ADD COLUMN category TEXT DEFAULT 'General'")
                    except: pass
                    conn.commit()
                    conn.close()
                except: pass
            
            ensure_schema()

            # --- ACTIONS ---
            def load_products():
                products_table.rows.clear()
                try:
                    conn = DBHandler.get_connection()
                    cursor = conn.cursor()
                    
                    cursor.execute("SELECT * FROM products ORDER BY rowid DESC")
                    rows = cursor.fetchall()
                    conn.close()
                    
                    query = search_tf.value.lower() if search_tf.value else ""
                    limit = 100 if not query else len(rows)
                    count = 0
                    
                    for row in rows:
                        if query and query not in row[1].lower(): continue
                        
                        # Handle old rows without expiry (row length check)
                        # Handle old rows (category is index 5 if viewing all columns, but fetched * is: barcode, name, price, stock, expiry, category)
                        # Actually sqlite * order depends on schema content. 
                        # We should be careful. Let's assume schema: barcode, name, price, stock, expiry_date, category. 
                        # But wait, original was: barcode, name, price, stock. 
                        # Then added expiry_date. Then added category.
                        # So: 0:barcode, 1:name, 2:price, 3:stock, 4:expiry (maybe), 5:category (maybe)
                        
                        cat = row[5] if len(row) > 5 else "General"
                        exp_date = row[4] if len(row) > 4 else "-"
                        if not exp_date: exp_date = "-"

                        products_table.rows.append(
                            ft.DataRow(
                                cells=[
                                    ft.DataCell(ft.Text(row[0])),
                                    ft.DataCell(ft.Text(row[1], weight="bold")),
                                    ft.DataCell(ft.Container(
                                        content=ft.Text(cat, size=12, color="white"),
                                        bgcolor="blue400", padding=5, border_radius=5
                                    )),
                                    ft.DataCell(ft.Text(f"{currency}{row[2]:.2f}")),
                                    ft.DataCell(ft.Text(f"{row[3]}")),
                                    ft.DataCell(ft.Text(exp_date, color="red" if exp_date != "-" else "black")),
                                    ft.DataCell(ft.Row([
                                        ft.IconButton(icon="edit", icon_color="blue", on_click=lambda e, r=row: start_edit(r)),
                                        ft.IconButton(icon="delete", icon_color="red", on_click=lambda e, code=row[0]: delete_product(code))
                                    ])),
                                ]
                            )
                        )
                        count += 1
                        if count >= limit: break
                        
                    self.page.update()
                except Exception as e:
                    print(e)
                    self.show_snack(f"Error loading: {e}", "red")

            def start_edit(row):
                edit_mode["active"] = True
                barcode_tf.value = row[0]
                barcode_tf.read_only = True 
                name_tf.value = row[1]
                price_tf.value = str(row[2])
                stock_tf.value = str(row[3])
                # Check for expiry & category
                expiry_tf.value = row[4] if len(row) > 4 and row[4] else ""
                category_tf.value = row[5] if len(row) > 5 and row[5] else "General"
                
                add_btn.text = "UPDATE ITEM"
                add_btn.bgcolor = "blue"
                self.page.update()

            def save_product(e):
                if not all([barcode_tf.value, name_tf.value, price_tf.value, stock_tf.value]):
                    self.show_snack("Please fill all fields (Barcode, Name, Price, Stock)", "red")
                    return
                try:
                    try:
                        p_val = float(price_tf.value)
                        s_val = int(stock_tf.value)
                    except ValueError:
                        self.show_snack("Price must be a number, Stock must be an integer", "red")
                        return

                    conn = DBHandler.get_connection()
                    cursor = conn.cursor()
                    
                    if edit_mode["active"]:
                        cursor.execute("UPDATE products SET name=?, category=?, price=?, stock=?, expiry_date=? WHERE barcode=?",
                                       (name_tf.value, category_tf.value, float(price_tf.value), int(stock_tf.value), expiry_tf.value, barcode_tf.value))
                        self.log_stock_change(name_tf.value, int(stock_tf.value), "Manager Update", cursor=cursor)
                        msg = "✅ Updated Successfully"
                        edit_mode["active"] = False
                        barcode_tf.read_only = False
                        add_btn.text = "ADD ITEM"
                        add_btn.bgcolor = "green"
                    else:
                        cursor.execute("INSERT INTO products (barcode, name, category, price, stock, expiry_date) VALUES (?, ?, ?, ?, ?, ?)", 
                                       (barcode_tf.value, name_tf.value, category_tf.value, float(price_tf.value), int(stock_tf.value), expiry_tf.value))
                        self.log_stock_change(name_tf.value, int(stock_tf.value), "Initial Stock", cursor=cursor)
                        msg = "✅ Added Successfully"

                    conn.commit()
                    conn.close()
                    
                    # Clear inputs
                    barcode_tf.value = ""
                    name_tf.value = ""
                    price_tf.value = ""
                    stock_tf.value = ""
                    expiry_tf.value = "" 
                    category_tf.value = ""
                    load_products()
                    self.show_snack(msg)
                except Exception as ex:
                    self.show_snack(f"Error: {ex}", "red")

            def delete_product(barcode):
                try:
                    conn = DBHandler.get_connection()
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM products WHERE barcode=?", (barcode,))
                    conn.commit()
                    conn.close()
                    load_products()
                    self.show_snack("🗑️ Item Deleted")
                except Exception as ex:
                    self.show_snack(f"Error: {ex}", "red")

            add_btn = ft.ElevatedButton("ADD ITEM", on_click=save_product, bgcolor="green", color="white", height=50)

            # --- LAYOUT ---
            bulk_actions = ft.Row([
                ft.Text("Bulk Actions:", weight="bold", size=14, color="grey"),
                ft.ElevatedButton("Download Template", icon="download", on_click=download_template, bgcolor="white", color="blue", height=30, style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=5), padding=5)),
                ft.ElevatedButton("Import Excel / CSV", icon="upload", on_click=lambda _: file_picker.pick_files(allowed_extensions=["xlsx", "xls", "csv"]), bgcolor="blue", color="white", height=30, style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=5), padding=5))
            ], spacing=10, alignment="end")

            self.page.add(
                ft.Container(
                    content=ft.Column([
                        ft.Text("Inventory Control", size=24, weight="bold", color=PRIMARY_COLOR),
                        # Form
                        ft.Container(
                            content=ft.Column([
                                ft.Text("Add / Edit Item", size=16, weight="bold", color="grey700"), 
                                ft.Row([barcode_tf, name_tf], spacing=10),
                                ft.Row([category_tf, price_tf, stock_tf, expiry_tf], spacing=10),
                                add_btn
                            ], spacing=10),
                            padding=20, bgcolor="white", border_radius=10, shadow=ft.BoxShadow(blur_radius=5, color="grey300")
                        ),
                        ft.Container(height=20),
                        
                        # List Header
                        ft.Row([
                            ft.Text("Stock List", size=20, weight="bold", color=PRIMARY_COLOR),
                            search_tf
                        ], alignment="spaceBetween"),
                        
                        bulk_actions,

                        ft.Container(
                            content=ft.Column([products_table], expand=True, scroll="auto"),
                            expand=True,
                            border=ft.border.all(1, "grey100"),
                            border_radius=5
                        )
                    ], spacing=10, expand=True),
                    padding=20,
                    expand=True 
                )
            )
            load_products()

        except Exception as e:
            import traceback
            traceback.print_exc()
            self.page.clean()
            self.page.add(ft.Text(f"Inventory Error: {e}", color="red", size=20))
            self.page.update()

    # 5. CRM (Customer Relationship Management)
    def show_crm(self):
        try:
            self.page.clean()
            self.set_app_bar("CRM & Client Management", show_back_button=True)
            
            # Inputs
            name_tf = ft.TextField(label="Full Name", prefix_icon="person")
            phone_tf = ft.TextField(label="Phone Number", prefix_icon="phone", width=200)
            
            clients_table = ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text("ID")),
                    ft.DataColumn(ft.Text("Client Name")),
                    ft.DataColumn(ft.Text("Phone")),
                    ft.DataColumn(ft.Text("Points/Loyalty")),
                    ft.DataColumn(ft.Text("Actions")),
                ],
                heading_row_color="grey200",
                border=ft.border.all(1, "grey300"),
                border_radius=10,
                column_spacing=20,
                width=float("inf") # Allow horizontal scroll if needed
            )

            def load_clients():
                clients_table.rows.clear()
                conn = DBHandler.get_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM customers ORDER BY id DESC")
                rows = cursor.fetchall()
                conn.close()
                
                for row in rows:
                    if row[1] == "Walk-in Client": continue
                    clients_table.rows.append(
                        ft.DataRow(cells=[
                            ft.DataCell(ft.Text(str(row[0]))),
                            ft.DataCell(ft.Text(row[1], weight="bold")),
                            ft.DataCell(ft.Text(row[2])),
                            ft.DataCell(ft.Container(content=ft.Text(str(row[3]), color="white", size=12), bgcolor="blue", padding=5, border_radius=5)),
                            ft.DataCell(ft.IconButton(icon="delete", icon_color="red", on_click=lambda e, uid=row[0]: delete_client(uid)))
                        ])
                    )
                self.page.update()

            def add_client(e):
                if not name_tf.value or not phone_tf.value:
                    self.show_snack("Please fill name and phone", "red")
                    return
                try:
                    conn = DBHandler.get_connection()
                    cursor = conn.cursor()
                    cursor.execute("INSERT INTO customers (name, phone, points) VALUES (?, ?, 0)", (name_tf.value, phone_tf.value))
                    conn.commit()
                    conn.close()
                    name_tf.value = ""
                    phone_tf.value = ""
                    load_clients()
                    self.show_snack("✅ Client Added")
                except Exception as ex:
                    self.show_snack(f"Error (duplicate phone?): {ex}", "red")

            def delete_client(uid):
                conn = DBHandler.get_connection()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM customers WHERE id=?", (uid,))
                conn.commit()
                conn.close()
                load_clients()
                self.show_snack("🗑️ Client Deleted")

            self.page.add(
                ft.Container(
                    content=ft.Column([
                        ft.Container(
                            content=ft.Row([
                                name_tf, phone_tf, 
                                ft.ElevatedButton("ADD CLIENT", on_click=add_client, bgcolor="green", color="white", height=50)
                            ], alignment="start"),
                            padding=20, bgcolor="white", border_radius=10, shadow=ft.BoxShadow(blur_radius=5, color="grey300")
                        ),
                        ft.Container(height=20),
                        ft.Text("Client Database", size=20, weight="bold", color=PRIMARY_COLOR),
                        ft.Container(
                            content=ft.Column([clients_table], scroll=ft.ScrollMode.AUTO),
                            padding=20, bgcolor="white", border_radius=10, shadow=ft.BoxShadow(blur_radius=5, color="grey300"),
                            expand=True
                        )
                    ], expand=True),
                    padding=20, expand=True
                )
            )
            load_clients()

        except Exception as e:
            self.page.add(ft.Text(f"CRM Error: {e}", color="red"))


    # 5. STAFF MANAGEMENT
    def show_staff_screen(self):
        self.page.clean()
        self.set_app_bar("Staff & HR Manager", show_back_button=True)
        
        # --- UI COMPONENTS ---
        username_tf = ft.TextField(label="Username", expand=True)
        pin_tf = ft.TextField(label="PIN (4 digits)", width=150, password=True)
        role_dd = ft.Dropdown(label="Role", width=150, options=[ft.dropdown.Option("Manager"), ft.dropdown.Option("Cashier")], value="Cashier")
        
        staff_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
        attendance_list = ft.Column(spacing=5, scroll=ft.ScrollMode.AUTO, expand=True)
        
        # Payroll Table
        payroll_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Staff Name", weight="bold")),
                ft.DataColumn(ft.Text("Base Salary ($)")),
                ft.DataColumn(ft.Text("Comm. Rate (%)")),
                ft.DataColumn(ft.Text("This Month Sales")),
                ft.DataColumn(ft.Text("Commission")),
                ft.DataColumn(ft.Text("TOTAL PAY")),
                ft.DataColumn(ft.Text("Actions")),
            ],
            border=ft.border.all(1, "grey300"),
            border_radius=10,
            heading_row_color="grey100",
            vertical_lines=ft.border.BorderSide(1, "grey200"),
            horizontal_lines=ft.border.BorderSide(1, "grey200"),
            width=float("inf")
        )

        # --- LOGIC ---
        
        def save_payroll_settings(uid, salary_val, comm_val):
            try:
                s = float(salary_val)
                c = float(comm_val)
                conn = DBHandler.get_connection()
                # Ensure columns exist (migration check logic used below)
                cursor = conn.cursor()
                cursor.execute("UPDATE users SET base_salary=?, commission_rate=? WHERE id=?", (s, c, uid))
                conn.commit()
                conn.close()
                self.show_snack("✅ Saved HR Settings")
                load_payroll() # Refresh calculations
            except Exception as ex: 
                self.show_snack(f"Error: {ex}", "red")

        def load_payroll():
            payroll_table.rows.clear()
            
            conn = DBHandler.get_connection()
            # conn.row_factory = sqlite3.Row -> Not supported by psycopg2 easily, use dictionary cursor logic or simple index
            cursor = conn.cursor()
            cursor.execute("SELECT id, username, role, base_salary, commission_rate FROM users")
            users = cursor.fetchall()
            
            # 2. Get Sales Performance (This Month)
            now_pak = get_pak_time()
            month_str = now_pak.strftime("%Y-%m")
            
            for u in users:
                if u[1] == "Admin": continue
                
                # Sales for this user this month
                # Fix: Use tuple index u[1] for username
                # Fix: CAST date_created for Postgres
                cursor.execute(f"SELECT SUM(total_amount) FROM offline_sales WHERE cashier_name=? AND CAST(date_created AS TEXT) LIKE '{month_str}%'", (u[1],))
                res = cursor.fetchone()
                total_sales = res[0] if res and res[0] else 0.0
                
                # Stats
                # Users tuple: id (0), username (1), role (2), base_salary (3), commission_rate (4)
                base = u[3] if u[3] else 0.0
                rate = u[4] if u[4] else 0.0
                
                commission_amt = total_sales * (rate / 100)
                total_pay = base + commission_amt
                
                # Inputs
                sal_field = ft.TextField(value=str(base), width=80, text_size=12, content_padding=10, suffix_text="$")
                comm_field = ft.TextField(value=str(rate), width=80, text_size=12, content_padding=10, suffix_text="%")
                
                # Save Button
                save_btn = ft.IconButton(
                    icon="save", 
                    icon_color="green", 
                    tooltip="Save Settings",
                    on_click=lambda e, uid=u[0], sf=sal_field, cf=comm_field: save_payroll_settings(uid, sf.value, cf.value)
                )

                payroll_table.rows.append(
                    ft.DataRow(cells=[
                        ft.DataCell(ft.Text(u["username"], weight="bold")),
                        ft.DataCell(sal_field),
                        ft.DataCell(comm_field),
                        ft.DataCell(ft.Text(f"${total_sales:,.2f}")),
                        ft.DataCell(ft.Text(f"${commission_amt:,.2f}", color="green")),
                        ft.DataCell(ft.Text(f"${total_pay:,.2f}", weight="bold", size=14, color=PRIMARY_COLOR)),
                        ft.DataCell(save_btn)
                    ])
                )
            conn.close()
            payroll_table.update()

        def load_staff():
            staff_list.controls.clear()
            conn = DBHandler.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id, username, role FROM users")
            rows = cursor.fetchall()
            conn.close()
            
            for row in rows:
                if row[1] == "Admin": continue 
                staff_list.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Icon("person", color=PRIMARY_COLOR),
                            ft.Text(row[1], expand=True, weight=ft.FontWeight.BOLD),
                            ft.Text(row[2], width=100),
                            ft.IconButton(icon="delete", icon_color="red", on_click=lambda e, uid=row[0]: delete_staff(uid))
                        ]),
                        bgcolor="white", padding=10, border_radius=5, border=ft.border.all(1, "grey200")
                    )
                )
            
            # Refresh attendance too
            load_attendance()
            # Refresh payroll too
            load_payroll()

        def load_attendance():
            attendance_list.controls.clear()
            conn = DBHandler.get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("SELECT user_name, action, timestamp FROM attendance ORDER BY id DESC LIMIT 50")
                logs = cursor.fetchall()
                if not logs: attendance_list.controls.append(ft.Text("No attendance records.", color="grey"))
                for log in logs:
                    color = "green" if log[1] == "IN" else "orange"
                    attendance_list.controls.append(
                        ft.Container(
                            content=ft.Row([
                                ft.Icon("access_time", size=16, color=color),
                                ft.Text(f"{log[0]}", weight="bold", width=100),
                                ft.Container(content=ft.Text(log[1], color="white", size=12), bgcolor=color, padding=5, border_radius=5),
                                ft.Text(log[2], size=12, color="grey")
                            ]),
                            bgcolor="white", padding=5, border_radius=5
                        )
                    )
            except: pass
            conn.close()

        def add_staff(e):
            if not all([username_tf.value, pin_tf.value, role_dd.value]):
                self.show_snack("Fill all fields", "red")
                return
            try:
                conn = DBHandler.get_connection()
                cursor = conn.cursor()
                # Init with 0 salary
                cursor.execute("INSERT INTO users (username, pin, role, base_salary, commission_rate) VALUES (?, ?, ?, 0, 0)", 
                               (username_tf.value, pin_tf.value, role_dd.value))
                conn.commit()
                conn.close()
                username_tf.value, pin_tf.value = "", ""
                load_staff()
                self.show_snack("✅ Staff Added")
            except Exception as ex:
                self.show_snack(f"Error: {ex}", "red")

        def delete_staff(user_id):
            try:
                conn = DBHandler.get_connection()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM users WHERE id=?", (user_id,))
                conn.commit()
                conn.close()
                load_staff()
                self.show_snack("🗑️ Staff Removed")
            except Exception as ex:
                self.show_snack(f"Error: {ex}", "red")

        # --- SETUP & TABS ---
        
        # 1. Ensure Columns Exist (Migration)
        try:
            conn = DBHandler.get_connection()
            cursor = conn.cursor()
            # Try add columns individually
            try: cursor.execute("ALTER TABLE users ADD COLUMN base_salary REAL DEFAULT 0") 
            except: pass
            try: cursor.execute("ALTER TABLE users ADD COLUMN commission_rate REAL DEFAULT 0")
            except: pass
            conn.commit()
            conn.close()
        except: pass

        # 2. Tabs
        tabs = ft.Tabs(
            selected_index=0,
            animation_duration=300,
            tabs=[
                ft.Tab(
                    text="Staff & Roles",
                    icon="people",
                    content=ft.Container(
                        content=ft.Row([
                            ft.Container(
                                content=ft.Column([
                                    ft.Container(
                                        content=ft.Row([
                                            username_tf, pin_tf, role_dd,
                                            ft.ElevatedButton("ADD STAFF", on_click=add_staff, bgcolor="green", color="white")
                                        ]),
                                        padding=20, bgcolor="white", border_radius=10, shadow=ft.BoxShadow(blur_radius=5, color="grey300")
                                    ),
                                    ft.Container(height=10),
                                    staff_list
                                ], expand=True),
                                expand=True, padding=20
                            ),
                            ft.Container(
                                content=ft.Column([
                                    ft.Text("Attendance Log", size=20, weight="bold", color=PRIMARY_COLOR),
                                    attendance_list
                                ], expand=True),
                                expand=True, padding=20, bgcolor="grey50", border_radius=10
                            )
                        ], expand=True),
                        padding=10
                    )
                ),
                ft.Tab(
                    text="Payroll Calculator",
                    icon="attach_money",
                    content=ft.Container(
                        padding=20,
                        content=ft.Column([
                            ft.Text("Payroll & Commission Dashboard", size=24, weight="bold", color=PRIMARY_COLOR),
                            ft.Text("Set Base Salary and Commission % per staff. Payouts update automatically based on this month's sales.", color="grey"),
                            ft.Divider(),
                            ft.Container(
                                content=ft.Column([payroll_table], scroll=ft.ScrollMode.AUTO),
                                expand=True,
                                bgcolor="white", padding=10, border_radius=10, border=ft.border.all(1, "grey200")
                            )
                        ], expand=True)
                    )
                )
            ],
            expand=True
        )

        self.page.add(tabs)
        load_staff()

    # 7. LABEL MAKER
    def show_label_maker(self):
        self.page.clean()
        self.set_app_bar("Barcode Label Creator", show_back_button=True)
        
        search_tf = ft.TextField(label="Find Product...", prefix_icon="search", expand=True)
        qty_slider = ft.Slider(min=1, max=50, divisions=49, label="{value} Stickers", value=10)
        
        # Selected Item State
        selected_item = {"barcode": "", "name": "", "price": 0.0}
        preview_container = ft.Container(content=ft.Text("Select a product to preview label"), alignment=ft.alignment.center, height=200, bgcolor="grey100", border_radius=10)
        
        def update_preview():
            if not selected_item["barcode"]:
                preview_container.content = ft.Text("Select a product first")
            else:
                # Show Mock Preview
                b_code = selected_item["barcode"]
                p_name = selected_item["name"]
                price = selected_item["price"]
                
                # We use the API url directly for preview Image
                url = f"https://bwipjs-api.metafloor.com/?bcid=code128&text={b_code}&scale=3&height=10&includetext"
                
                preview_container.content = ft.Column([
                    ft.Text(p_name, weight="bold", size=16),
                    ft.Image(src=url, height=80, fit=ft.ImageFit.CONTAIN),
                    ft.Text(f"${price:.2f}", weight="bold", size=14),
                    ft.Text(f"Printing {int(qty_slider.value)} copies", color="blue")
                ], alignment="center", horizontal_alignment="center")
            self.page.update()

        def select_product(row):
            selected_item["barcode"] = row[0]
            selected_item["name"] = row[1]
            selected_item["price"] = row[2]
            search_tf.value = row[1]
            results_list.controls.clear()
            update_preview()
            
        results_list = ft.Column(scroll="auto", height=200)

        def run_search(e):
            term = search_tf.value.lower()
            results_list.controls.clear()
            if not term: 
                self.page.update()
                return
            
            conn = DBHandler.get_connection()
            cursor = conn.cursor()
            # Postgres LIKE is case sensitive, use ILIKE? or just lower() wrapper.
            # For compatibility:
            if DBHandler.HAS_POSTGRES:
                 cursor.execute("SELECT barcode, name, price FROM products WHERE LOWER(name) LIKE ? OR CAST(barcode AS TEXT) LIKE ? LIMIT 10", (f"%{term}%", f"%{term}%"))
            else:
                 cursor.execute("SELECT barcode, name, price FROM products WHERE name LIKE ? OR barcode LIKE ? LIMIT 10", (f"%{term}%", f"%{term}%"))
            rows = cursor.fetchall()
            conn.close()
            
            for r in rows:
                results_list.controls.append(
                    ft.ListTile(
                        title=ft.Text(r[1]),
                        subtitle=ft.Text(f"{r[0]} - ${r[2]}"),
                        on_click=lambda e, row=r: select_product(row),
                        bgcolor="white"
                    )
                )
            self.page.update()

        search_tf.on_change = run_search
        qty_slider.on_change = lambda e: update_preview()

        def print_labels(e):
            if not selected_item["barcode"]:
                self.show_snack("Select a product first!", "red")
                return
            
            if not generate_barcode_sheet:
                self.show_snack("❌ PDF Module not available", "red")
                return

            self.show_snack("Generating PDF...", "blue")
            try:
                fname = generate_barcode_sheet(selected_item["barcode"], selected_item["name"], selected_item["price"], int(qty_slider.value))
                if fname.startswith("Error"): self.show_snack(fname, "red")
                else: self.show_snack(f"✅ Labels Ready: {fname}")
            except Exception as ex:
                self.show_snack(f"Error: {ex}", "red")

        self.page.add(
            ft.Container(
                content=ft.Column([
                    ft.Text("Step 1: Select Product", size=16, weight="bold", color=PRIMARY_COLOR),
                    search_tf,
                    results_list,
                    ft.Divider(),
                    ft.Text("Step 2: Configure Labels", size=16, weight="bold", color=PRIMARY_COLOR),
                    ft.Row([ft.Text("Quantity:"), qty_slider], alignment="spaceBetween"),
                    preview_container,
                    ft.Container(height=10),
                    ft.ElevatedButton("PRINT LABELS (PDF)", on_click=print_labels, bgcolor="purple", color="white", height=50, width=300)
                ], scroll="auto", expand=True),
                padding=20, bgcolor="white", border_radius=10, shadow=ft.BoxShadow(blur_radius=5, color="grey300"),
                margin=20, expand=True
            )
        )

    # 6. SETTINGS SCREEN
    def show_settings_screen(self):
        self.page.clean()
        self.set_app_bar("System Settings", show_back_button=True)
        
        shop_name_tf = ft.TextField(label="Shop Name", value=self.get_setting('shop_name'))
        address_tf = ft.TextField(label="Address", value=self.get_setting('address'))
        phone_tf = ft.TextField(label="Phone Info", value=self.get_setting('phone'))
        currency_tf = ft.TextField(label="Currency Symbol", value=self.get_setting('currency'), width=100)
        
        def save_settings(e):
            self.set_setting('shop_name', shop_name_tf.value)
            self.set_setting('address', address_tf.value)
            self.set_setting('phone', phone_tf.value)
            self.set_setting('currency', currency_tf.value)
            self.show_snack("✅ Settings Saved! Return to Dashboard to see changes.")

        self.page.add(
            ft.Container(
                content=ft.Column([
                    ft.Text("Shop Configuration", size=24, weight=ft.FontWeight.BOLD, color=PRIMARY_COLOR),
                    ft.Container(height=20),
                    shop_name_tf,
                    address_tf,
                    phone_tf,
                    phone_tf,
                    currency_tf,
                    # ft.Switch(label="Dark Mode", on_change=self.toggle_theme, value=self.page.theme_mode == ft.ThemeMode.DARK),
                    ft.Container(height=20),
                    ft.ElevatedButton("SAVE CHANGES", on_click=save_settings, bgcolor=PRIMARY_COLOR, color="white", height=50, width=200)
                ], width=600),
                alignment=ft.alignment.center, padding=20
            )
        )

def main(page: ft.Page):
    app = HornERP(page)

if __name__ == "__main__":
    # Render provides PORT environment variable
    port = int(os.environ.get("PORT", 8550))
    print(f"\n[STARTING] HORN ERP...")
    print(f"[LOCAL] Access: http://127.0.0.1:{port}")
    print(f"[CLOUD] Render Port: {port} (Listening on 0.0.0.0)\n")
    
    ft.app(target=main, view=ft.WEB_BROWSER, port=port, host="0.0.0.0", assets_dir="assets")
