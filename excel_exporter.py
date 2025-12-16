import datetime
import os
try:
    from db_connector import DBHandler
except ImportError:
    # Fallback if running standalone
    DBHandler = None
    import sqlite3

def export_to_excel(period="Daily", is_native=False):
    """
    Exports data to a real Excel file with multiple sheets if openpyxl is available.
    Otherwise falls back to a basic HTML-Excel dump or CSV.
    Uses DBHandler to ensure data comes from the active database (SQLite or Postgres).
    Args:
        is_native (bool): If True, saves to Desktop (Local App). If False, saves to assets/reports (Web App).
    """
    # Try importing openpyxl
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        HAS_OPENPYXL = True
    except ImportError:
        HAS_OPENPYXL = False

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    if is_native:
        # Native Mode
        reports_dir = os.path.join(os.environ['USERPROFILE'], 'Desktop')
        if not os.path.exists(reports_dir): os.makedirs(reports_dir) # Just in case
    else:
        # Web Mode
        base_dir = os.path.dirname(os.path.abspath(__file__))
        reports_dir = os.path.join(base_dir, "assets", "reports")
        if not os.path.exists(reports_dir):
            os.makedirs(reports_dir)
    
    # Connect
    if DBHandler:
        conn = DBHandler.get_connection()
        # Determine placeholder style
        ph = DBHandler.get_placeholder() 
    else:
        conn = sqlite3.connect("horn.db")
        ph = "?"

    if not conn:
        return None
        
    cursor = conn.cursor()

    # --- TIMEZONE FIX ---
    # Need to be careful with Postgres vs SQLite generic SQL if possible
    # For filtering, Python-side date generation usually works best to form string literals
    
    tz_offset = datetime.timezone(datetime.timedelta(hours=5))
    now_pak = datetime.datetime.now(tz_offset)
    today_str = now_pak.strftime("%Y-%m-%d")
    month_str = now_pak.strftime("%Y-%m")
    
    # Filter Logic
    # Postgres uses "TEXT" casting sometimes for LIKE, or we can use specific function
    # DBHandler usually connects to Postgres, but standard ANSI SQL is preferred.
    
    where_clause = ""
    
    # We will use simple string matching which works enough for YYYY-MM-DD strings in DB
    # Note: On Postgres, timestamps might need casting to text for LIKE to work effortlessly on left side
    
    date_col = "s.date_created"
    # Safe cast helper
    date_col_cast = f"CAST({date_col} AS TEXT)" # Works in PG and SQLite usually
    
    if period == "Daily":
        where_clause = f"{date_col_cast} LIKE '{today_str}%'"
    elif period == "Weekly":
        start_date = (now_pak - datetime.timedelta(days=7)).strftime("%Y-%m-%d")
        # For weekly, simple string comparison YYYY-MM-DD works
        where_clause = f"{date_col_cast} >= '{start_date}'"
    elif period == "Monthly":
        where_clause = f"{date_col_cast} LIKE '{month_str}%'"
    else:
        where_clause = f"{date_col_cast} LIKE '{today_str}%'"

    # 1. Fetch Inventory
    # Handle schema differences via try/except or generic select is risky
    # Let's try to select specific columns. 
    # NOTE: Postgres is case sensitive on column names if quoted, but generally fine unquoted.
    
    try:
        cursor.execute("SELECT barcode, name, category, price, stock, expiry_date FROM products ORDER BY category, name")
    except:
        # Fallback if 'category' or 'expiry_date' doesn't exist (though migration should have fixed it)
        conn.rollback() # Postgres requires rollback after error
        cursor.execute("SELECT barcode, name, 'General', price, stock, '' FROM products ORDER BY name")
        
    inventory_data = cursor.fetchall()

    # 2. Fetch Sales Detailed (Gross)
    # COALESCE is standard SQL
    query = f"""
        SELECT 
            s.date_created, 
            s.cashier_name, 
            s.customer_name,
            COALESCE(p.category, 'General') as category,
            i.product_name,
            i.quantity,
            i.price,
            (i.quantity * i.price) as line_total
        FROM offline_sale_items i
        JOIN offline_sales s ON i.sale_id = s.id
        LEFT JOIN products p ON i.product_name = p.name
        WHERE {where_clause}
        ORDER BY s.date_created DESC
    """
    try:
        cursor.execute(query)
        sales_data = cursor.fetchall()
    except Exception as e:
        print(f"Excel Export Query Error: {e}")
        conn.rollback()
        sales_data = []
    
    # 3. Net Revenue Calculation
    try:
        # Re-use clause but targeting table alias 's' (which is offline_sales s)
        # Wait, simple fetch from sales table
        simple_clause = where_clause.replace("s.date_created", "date_created")
        cursor.execute(f"SELECT SUM(total_amount) FROM offline_sales WHERE {simple_clause}")
        res = cursor.fetchone()
        net_revenue = res[0] if res and res[0] else 0.0
    except:
        net_revenue = 0.0
    
    conn.close()

    # --- GENERATION ---
    filename = f"HornERP_Report_{timestamp}.xlsx" if HAS_OPENPYXL else f"HornERP_Report_{timestamp}.xls"
    filepath = os.path.join(reports_dir, filename)
    web_path = f"/reports/{filename}" # Relative path for Flet

    if HAS_OPENPYXL:
        # --- MODE A: Real Excel (.xlsx) ---
        wb = openpyxl.Workbook()
        
        # Sheet 1: Sales
        ws_sales = wb.active
        ws_sales.title = "Sales Report"
        headers = ["Date", "Cashier", "Customer", "Category", "Product", "Qty", "Price", "Total (Gross)"]
        ws_sales.append(headers)
        
        # Style Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        for cell in ws_sales[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        gross_revenue = 0
        for row in sales_data:
            ws_sales.append(list(row))
            gross_revenue += float(row[7] or 0)
            
        # Summary
        discount_total = gross_revenue - float(net_revenue)
        ws_sales.append([])
        ws_sales.append(["", "", "", "", "", "", "Gross Sales:", gross_revenue])
        ws_sales.append(["", "", "", "", "", "", "Discounts:", -discount_total])
        ws_sales.append(["", "", "", "", "", "", "NET REVENUE:", net_revenue])
        
        # Formatting Summary
        lr = ws_sales.max_row
        ws_sales[f"H{lr}"].font = Font(bold=True, color="0000FF")
        
        ws_sales.column_dimensions['A'].width = 20
        ws_sales.column_dimensions['E'].width = 25

        # Sheet 2: Inventory
        ws_inv = wb.create_sheet("Inventory Status")
        ws_inv.append(["Barcode", "Name", "Category", "Price", "Stock", "Expiry"])
        for cell in ws_inv[1]:
            cell.font = header_font
            cell.fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
        
        for row in inventory_data:
            ws_inv.append(list(row))
            
        ws_inv.column_dimensions['B'].width = 30
        
        wb.save(filepath)
    else:
        # --- MODE B: HTML Fallback ---
        html_content = f"""
        <html><body>
        <h1>Horn ERP Report ({period})</h1>
        <h2>Sales</h2>
        <table border='1'>
        <tr><th>Date</th><th>Product</th><th>Qty</th><th>Price</th><th>Total</th></tr>
        {"".join(f"<tr><td>{r[0]}</td><td>{r[4]}</td><td>{r[5]}</td><td>{r[6]}</td><td>{r[7]}</td></tr>" for r in sales_data)}
        </table>
        <h2>Inventory</h2>
        <table border='1'>
        <tr><th>Barcode</th><th>Name</th><th>Stock</th></tr>
        {"".join(f"<tr><td>{r[0]}</td><td>{r[1]}</td><td>{r[4]}</td></tr>" for r in inventory_data)}
        </table>
        </body></html>
        """
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html_content)

    print(f"Generated Excel: {filepath}")
    return filepath if is_native else filename

